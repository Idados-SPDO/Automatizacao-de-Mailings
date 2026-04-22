import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import data_processing as dp
import datetime as dt
from io import BytesIO


def page_carrega_dado():
    st.header('Ferramenta de :red[Automatização de Mailings]', divider="red")
    st.write('')
    st.write('')

    container = st.container()
    container.subheader('Digite a data de Referência para preenchimento das planilhas:')
    data_referencia = container.date_input("", value=None, format="DD/MM/YYYY")

    if data_referencia is not None:
        container.write(f'Data selecionada: {dp.dia(data_referencia)}/{dp.mes(data_referencia)}/{dp.ano(data_referencia)}')

        container.subheader('Importe os arquivos de relatório utilizados para o preenchimento das planilhas:')
        uploaded_files = container.file_uploader('', accept_multiple_files=True, type=["xlsx"])

        with st.spinner('Preparando e guardando os dados de preenchimento. Por favor aguarde...'):
            st.session_state.update({"data_referencia": data_referencia})

            if len(uploaded_files) >= 1:
                for uploaded_file in uploaded_files:
                    if uploaded_file.name.startswith('ECE_'):
                        st.session_state.update({"Dados_status_ECE": dp.load_mailing_status(uploaded_file)})
                    if uploaded_file.name.startswith('ECMA_'):
                        st.session_state.update({"Dados_status_ECMA": dp.load_mailing_status(uploaded_file)})
                    if uploaded_file.name.startswith('ECI_'):
                        st.session_state.update({"Dados_status_ECI": dp.load_mailing_status(uploaded_file)})
                    if uploaded_file.name.startswith('SCM_'):
                        st.session_state.update({"Dados_geral_SCM": dp.load_mailing_geral(uploaded_file)})
                        st.session_state.update({"Dados_status_SCM": dp.load_mailing_status(uploaded_file)})
                        st.session_state.update({"Dados_placar_SCM": dp.load_mailing_placar(uploaded_file)})
                    if uploaded_file.name.startswith('SCC_'):
                        st.session_state.update({"Dados_geral_SCC": dp.load_mailing_geral(uploaded_file)})
                        st.session_state.update({"Dados_status_SCC": dp.load_mailing_status(uploaded_file)})
                        st.session_state.update({"Dados_placar_SCC": dp.load_mailing_placar(uploaded_file)})
                    if uploaded_file.name.startswith('SC_'):
                        st.session_state.update({"Dados_geral_SC": dp.load_mailing_geral(uploaded_file)})
                        st.session_state.update({"Dados_status_SC": dp.load_mailing_status(uploaded_file)})
                        st.session_state.update({"Dados_placar_SC": dp.load_mailing_placar(uploaded_file)})
                    if uploaded_file.name.startswith('SSV_'):
                        st.session_state.update({"Dados_geral_SSV": dp.load_mailing_geral(uploaded_file)})
                        st.session_state.update({"Dados_status_SSV": dp.load_mailing_status(uploaded_file)})
                        st.session_state.update({"Dados_placar_SSV": dp.load_mailing_placar(uploaded_file)})

                st.success('Arquivos importados com sucesso!')


def page_preenche_indicador():
    sondagem_map = [
        ("Dados_status_ECE", "Consumidor (E)"),
        ("Dados_status_ECMA", "Consumidor (MA)"),
        ("Dados_status_ECI", "Consumidor (I)"),
        ("Dados_status_SCM", "Comércio"),
        ("Dados_status_SCC", "Construção"),
        ("Dados_status_SC", "Indústria"),
        ("Dados_status_SSV", "Serviços"),
    ]

    arquivos_inseridos = []
    sondagens_trabalhadas = []
    for chave, nome in sondagem_map:
        dado = st.session_state.get(chave)
        if dado is not None:
            arquivos_inseridos.append(dado)
            sondagens_trabalhadas.append(nome)

    if arquivos_inseridos:
        st.header('Preencher :red[Indicador Status Sondagem]', divider="red")
        st.write('')
        st.write('')

        container = st.container()
        container.subheader('Importe arquivo a ser atualizado:')
        uploaded_indicador = container.file_uploader('', accept_multiple_files=False, type=["xlsx"], key="page1")

        if uploaded_indicador is not None:
            workbook = load_workbook(uploaded_indicador)

            for i in range(len(sondagens_trabalhadas)):
                workbook = dp.preenche_indicador(arquivos_inseridos[i], workbook, sondagens_trabalhadas[i])

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            st.write('')
            st.write('')

            st.download_button(
                label="Baixar arquivo atualizado",
                data=buffer,
                file_name=f'Indicador_STATUS_Sondagem - Atualizado_{dp.dia(dt.date.today())}/{dp.mes(dt.date.today())}.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning('Por favor insira as Planilhas em "Importar planilha base".')


def page_preenche_status():
    if st.session_state.Dados_status_EC is None:
        df_ECE = st.session_state.Dados_status_ECE
        df_ECMA = st.session_state.Dados_status_ECMA
        df_ECI = st.session_state.Dados_status_ECI

        df_EC = df_ECE.copy()
        df_EC['QUANTIDADE'] += df_ECMA['QUANTIDADE'] + df_ECI['QUANTIDADE']
        st.session_state.update({"Dados_status_EC": df_EC})

    sondagem_map = [
        ("Dados_status_EC", "Consumidor"),
        ("Dados_status_SCM", "Comércio"),
        ("Dados_status_SCC", "Construção"),
        ("Dados_status_SC", "Indústria"),
        ("Dados_status_SSV", "Serviços"),
    ]

    arquivos_inseridos = []
    sondagens_trabalhadas = []
    for chave, nome in sondagem_map:
        dado = st.session_state.get(chave)
        if dado is not None:
            arquivos_inseridos.append(dado)
            sondagens_trabalhadas.append(nome)

    if arquivos_inseridos:
        st.header('Preencher :red[SACE Status Sondagem]', divider="red")
        st.write('')
        st.write('')

        container = st.container()
        container.subheader('Importe arquivo a ser atualizado:')
        uploaded_status = container.file_uploader('', accept_multiple_files=False, type=["xlsx"], key="page2")

        if uploaded_status is not None:
            workbook = load_workbook(uploaded_status)

            for i in range(len(sondagens_trabalhadas)):
                workbook = dp.preenche_status(arquivos_inseridos[i], workbook, sondagens_trabalhadas[i].upper())

            buffer = BytesIO()
            workbook.save(buffer)
            buffer.seek(0)

            st.dataframe(pd.read_excel(buffer), width=800, height=300)

            st.write('')
            st.write('')

            buffer.seek(0)
            st.download_button(
                label="Baixar arquivo atualizado",
                data=buffer,
                file_name=f'BD_SACE_Sondagem_Status_{dp.dia(dt.date.today())}.{dp.mes(dt.date.today())}.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning('Por favor insira as Planilhas em "Importar planilha base".')


def page_preenche_taxa_resposta():
    sondagem_map = [
        ("Dados_geral_SCM", "Comércio"),
        ("Dados_geral_SCC", "Construção"),
        ("Dados_geral_SC", "Indústria"),
        ("Dados_geral_SSV", "Serviços"),
    ]

    arquivos_inseridos = []
    sondagens_trabalhadas = []
    for chave, nome in sondagem_map:
        dado = st.session_state.get(chave)
        if dado is not None:
            arquivos_inseridos.append(dado)
            sondagens_trabalhadas.append(nome)

    if arquivos_inseridos:
        st.header('Preencher :red[SACE Sondagem Taxa de Resposta]', divider="red")
        st.write('')
        st.write('')

        container = st.container()
        container.subheader('Importe arquivo a ser atualizado:')
        uploaded_taxa = container.file_uploader('', accept_multiple_files=False, type=["xlsx"], key="page3")

        if uploaded_taxa is not None:
            df_uploaded = pd.read_excel(uploaded_taxa)

            for i in range(len(sondagens_trabalhadas)):
                df_uploaded = dp.preenche_taxa_resposta(arquivos_inseridos[i], df_uploaded, sondagens_trabalhadas[i])

            buffer = BytesIO()
            with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                df_uploaded.to_excel(writer, index=False, sheet_name="BD_SACE")
            buffer.seek(0)

            st.dataframe(pd.read_excel(buffer), width=800, height=300)

            st.write('')
            st.write('')

            buffer.seek(0)
            st.download_button(
                label="Baixar arquivo atualizado",
                data=buffer,
                file_name=f'BD_SACE_Sondagem_Taxa de Respostas_{dp.dia(dt.date.today())}.{dp.mes(dt.date.today())}.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning('Por favor insira as Planilhas em "Importar planilha base".')


def page_preenche_prioritarias():
    sondagem_map = [
        ("Dados_placar_SCM", "Comércio"),
        ("Dados_placar_SCC", "Construção"),
        ("Dados_placar_SC", "Indústria"),
        ("Dados_placar_SSV", "Serviços"),
    ]

    arquivos_inseridos = []
    sondagens_trabalhadas = []
    for chave, nome in sondagem_map:
        dado = st.session_state.get(chave)
        if dado is not None:
            arquivos_inseridos.append(dado)
            sondagens_trabalhadas.append(nome)

    if arquivos_inseridos:
        st.header('Preencher :red[SACE Sondagens Prioritárias]', divider="red")
        st.write('')
        st.write('')

        container = st.container()
        container.subheader('Importe arquivo a ser atualizado:')
        uploaded_prioritarias = container.file_uploader('', accept_multiple_files=False, type=["xlsx"], key="page4")

        if uploaded_prioritarias is not None:
            with st.spinner('Preenchendo a nova planilha. Por favor aguarde...'):
                workbook = load_workbook(uploaded_prioritarias)

                for i in range(len(sondagens_trabalhadas)):
                    workbook = dp.preenche_prioritarias(arquivos_inseridos[i], workbook, sondagens_trabalhadas[i])

                buffer = BytesIO()
                workbook.save(buffer)
                buffer.seek(0)

            st.dataframe(pd.read_excel(buffer), width=800, height=300)

            st.write('')
            st.write('')

            buffer.seek(0)
            st.download_button(
                label="Baixar arquivo atualizado",
                data=buffer,
                file_name=f'BD_SACE_Sondagens_Prioritárias_{dp.dia(dt.date.today())}.{dp.mes(dt.date.today())}.xlsx',
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
    else:
        st.warning('Por favor insira as Planilhas em "Importar planilha base".')
