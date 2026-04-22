import streamlit as st
from openpyxl import load_workbook
import pandas as pd
import datetime as dt
from babel.dates import format_date


def load_mailing_geral(content_file=None):
    if content_file is not None:
        return pd.read_excel(content_file, sheet_name="Mailing Geral")

def load_mailing_status(content_file=None):
    if content_file is not None:
        return pd.read_excel(content_file, sheet_name="Status")

def load_mailing_placar(content_file=None):
    if content_file is not None:
        data_preenchimento = st.session_state.data_referencia
        return pd.read_excel(
            content_file,
            sheet_name=f'Placar {dia(data_preenchimento)}.{mes(data_preenchimento)}',
            header=None
        )

def load_sondagem_status(content_file=None):
    if content_file is not None:
        content = pd.read_excel(content_file, sheet_name="Sondagem_Status")


def preenche_indicador(df_status, workbook, sheet):
    data_preenchimento = st.session_state.data_referencia
    sheet_workbook = workbook[sheet]

    for indice, linha in df_status.iterrows():
        for row in sheet_workbook.iter_rows(min_row=2):
            if linha['STATUS'].strip() in row[2].value:
                sheet_workbook.cell(row=row[0].row, column=1, value=f'{dia(data_preenchimento)}/{mes(data_preenchimento)}/{ano(data_preenchimento)}')
                sheet_workbook.cell(row=row[0].row, column=4, value=linha['QUANTIDADE'])
                sheet_workbook.cell(row=row[0].row, column=5, value=linha['%'])
                break

    return workbook

def preenche_status(df_status, workbook, sondagem):
    if 'CONSUMIDOR' in sondagem:
        status_to_group = {
            'CONCLUÍDA POR CC[1]': 'REALIZADO',
            'CONCLUÍDA POR E-MAIL / WPP[2]': 'REALIZADO',
            'AGENDAMENTO[3]': 'EM_NEGOCIACAO',
            'E-MAIL ENVIADO[4]': 'EM_NEGOCIACAO',
            'WHATSAPP ENVIADO[5]': 'EM_NEGOCIACAO',
            'NÃO ENCONTRADO[12]': 'EM_NEGOCIACAO',
            'AINDA NÃO TRABALHADO[7]': 'PENDENTE',
            'NÃO TEM WHATSAPP[8]': 'PENDENTE',
            'NÃO DESEJA MAIS PARTICIPAR[9]': 'DESATIVAR',
            'SEM PERFIL [10]': 'DESATIVAR',
            'RECUSA NO MÊS[11]': 'SEM_SUCESSO',
            'TELEFONE ERRADO[13]': 'SEM_SUCESSO',
            'TELEFONE OCUPADO[14]': 'SEM_SUCESSO',
            'NÃO ATENDE[15]': 'SEM_SUCESSO',
            'NÃO COMPLETA[16]': 'SEM_SUCESSO',
            'NÃO EXISTE[17]': 'SEM_SUCESSO',
            'SECRETÁRIA ELETRÔNICA/CAIXA POSTAL[18]': 'SEM_SUCESSO',
            'INFORMANTE OCUPADO[19]': 'SEM_SUCESSO',
            'VINCULADO JÁ POSSUI RESPOSTA[20]': 'SEM_SUCESSO',
            'NÃO LIGAR[21]': 'SEM_SUCESSO',
            'CONTATO REPETIDO[22]': 'SEM_SUCESSO',
            'INFORMANTE NÃO CADASTRADO[23]': 'SEM_SUCESSO',
            'NÚMERO INCOMPLETO[24]': 'SEM_SUCESSO',
        }
    else:
        status_to_group = {
            'CONCLUÍDA': 'REALIZADO',
            'AGENDAMENTO': 'EM_NEGOCIACAO',
            'E-MAIL ENVIADO': 'EM_NEGOCIACAO',
            'WHATSAPP ENVIADO': 'EM_NEGOCIACAO',
            'NÃO ENCONTRADO': 'EM_NEGOCIACAO',
            'EM NEGOCIAÇÃO': 'EM_NEGOCIACAO',
            'AINDA NÃO TRABALHADO': 'PENDENTE',
            'RECUSA NO MÊS': 'SEM_SUCESSO',
            'TELEFONE ERRADO': 'SEM_SUCESSO',
            'PROBLEMA NO TELEFONE': 'SEM_SUCESSO',
            'NÃO DESEJA MAIS PARTICIPAR': 'DESATIVAR',
            'SEM PERFIL': 'DESATIVAR',
            'EMPRESA FECHADA': 'EMP_FECHADA',
            'MAILING EXCEDENTE': 'BASE_EXTRA',
            'TRATAR COM RELACIONAMENTO': 'RECUPERAR',
            'EM ESTUDO': 'RECUPERAR',
            'EMPRESA RECUPERADA/ PROSPECTADA': 'PROSPECTADO/RECUPERADO',
        }

    df_status['GRUPO DE STATUS'] = df_status['STATUS'].str.strip().map(status_to_group).fillna('')
    df_status = df_status[['GRUPO DE STATUS', 'STATUS', 'QUANTIDADE', '%']]
    df_status = df_status.drop(columns=['STATUS', '%'])
    df_status = df_status[df_status['GRUPO DE STATUS'] != '']
    df_status = df_status.groupby('GRUPO DE STATUS', as_index=False).sum()

    sheet_workbook = workbook['BD_SACE_Sondagem_Status']
    data_atual = f'{dia(st.session_state.data_referencia)}/{mes(st.session_state.data_referencia)}/{ano(st.session_state.data_referencia)}'

    tamanho_sondagem = {
        'COMÉRCIO': 800,
        'CONSTRUÇÃO': 700,
        'INDÚSTRIA': 1150,
        'SERVIÇOS': 1700,
        'CONSUMIDOR': 2045,
    }

    if sondagem not in tamanho_sondagem:
        return workbook

    tamanho = tamanho_sondagem[sondagem]
    for indice, linha in df_status.iterrows():
        new_row = [data_atual, sondagem, tamanho, linha['GRUPO DE STATUS'], linha['QUANTIDADE']]
        sheet_workbook.append(new_row)

    return workbook

def preenche_taxa_resposta(df_taxa, df_uploaded, sondagem):
    df_taxa = df_taxa.dropna(subset=['Placar'])
    df_taxa['Placar'] = pd.to_datetime(df_taxa['Placar'], format="%d/%m/%Y")

    def determinar_acao(row):
        min_date = df_taxa['Placar'].min()
        data_minima = f'{dia(min_date)}/{mes(min_date)}/{ano(min_date)}'
        data_analise = f'{dia(row["Placar"])}/{mes(row["Placar"])}/{ano(row["Placar"])}'
        if pd.isnull(row['Placar']):
            return ''
        elif data_analise == data_minima:
            return 'espontânea'
        else:
            return 'esforço equipe'

    df_taxa['Ação'] = df_taxa.apply(determinar_acao, axis=1)

    novas_linhas = [
        {'Sondagem': sondagem, 'Porte': 'P', 'Ação': 'esforço equipe'},
        {'Sondagem': sondagem, 'Porte': 'M', 'Ação': 'esforço equipe'},
        {'Sondagem': sondagem, 'Porte': 'G', 'Ação': 'esforço equipe'},
        {'Sondagem': sondagem, 'Porte': 'P', 'Ação': 'espontânea'},
        {'Sondagem': sondagem, 'Porte': 'M', 'Ação': 'espontânea'},
        {'Sondagem': sondagem, 'Porte': 'G', 'Ação': 'espontânea'},
    ]

    df_uploaded = pd.concat([df_uploaded, pd.DataFrame(novas_linhas)], ignore_index=True)

    filtro_sondagem = (df_uploaded['Sondagem'] == sondagem) & (df_uploaded['Data'].isnull())
    data_atual = f'{dia(st.session_state.data_referencia)}/{mes(st.session_state.data_referencia)}/{ano(st.session_state.data_referencia)}'
    data_atual = pd.to_datetime(data_atual, format='%d/%m/%Y')

    df_taxa_resposta = df_uploaded[filtro_sondagem].copy()
    indices_taxa_resposta = df_taxa_resposta.index.tolist()

    for indice in indices_taxa_resposta:
        df_taxa_resposta.loc[indice, 'Data'] = data_atual.strftime('%d/%m/%Y')

    contagem_por_grupo = df_taxa.groupby(['Ação', 'Porte']).size().reset_index(name='Quantidade')
    df_taxa_resposta = pd.merge(df_taxa_resposta, contagem_por_grupo, on=['Ação', 'Porte'], how='left', suffixes=('_df_taxa_resposta', '_contagem_por_grupo'))
    df_taxa_resposta['Quantidade'] = df_taxa_resposta['Quantidade_contagem_por_grupo'].fillna(0).astype(int)
    df_taxa_resposta = df_taxa_resposta.drop(['Quantidade_df_taxa_resposta', 'Quantidade_contagem_por_grupo'], axis=1)
    df_taxa_resposta = df_taxa_resposta.drop_duplicates(subset=['Ação', 'Porte'])

    df_uploaded['Data'] = pd.to_datetime(df_uploaded['Data'], format="%d/%m/%Y").dt.strftime('%d/%m/%Y')

    filtro_sondagem = (df_uploaded['Sondagem'] == sondagem) & (
        (df_uploaded['Data'].isin(df_taxa_resposta['Data'])) | (df_uploaded['Data'].isnull())
    )
    linhas_para_atualizar = df_uploaded[filtro_sondagem].head(6)

    if not linhas_para_atualizar.empty:
        colunas_para_atualizar = ['Data', 'Quantidade']
        for x, (idx, row) in enumerate(linhas_para_atualizar.iterrows()):
            df_uploaded.loc[idx, colunas_para_atualizar] = df_taxa_resposta.loc[x, colunas_para_atualizar].values

    return df_uploaded

def preenche_prioritarias(df_placar, workbook, sondagem):
    sheet_workbook = workbook['Base']

    placar_renomeacao = {0: 'Código', 1: 'Nome', 2: 'Data', 3: 'Tipo'}
    df_placar = df_placar.rename(columns=placar_renomeacao)

    for row in sheet_workbook.iter_rows(min_row=2):
        if row[2].value in df_placar['Código'].values:
            i_corresp = df_placar[df_placar['Código'] == row[2].value].index[0]
            sheet_workbook.cell(row=row[0].row, column=1, value=df_placar['Data'][i_corresp])
            sheet_workbook.cell(row=row[0].row, column=2, value=sondagem)
            sheet_workbook.cell(row=row[0].row, column=6, value='sim')

    return workbook


def dia(data):
    return dt.datetime.strftime(data, '%d')

def mes(data):
    return dt.datetime.strftime(data, '%m')

def nome_mes(data):
    return format_date(data, 'MMMM', locale='pt_BR').capitalize()

def ano(data):
    return dt.datetime.strftime(data, '%Y')
